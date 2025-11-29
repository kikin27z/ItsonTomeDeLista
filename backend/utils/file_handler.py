# attendance/exporters.py
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from datetime import datetime
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


class AttendanceExporter:
    """Clase base para exportadores de asistencia"""

    STATUS_MAP = {
        'PRESENT': 'Presente',
        'LATE': 'Tardía',
        'ABSENT': 'Ausente',
        'JUSTIFIED': 'Justificada'
    }

    def __init__(self, data):
        """
        data: dict con 'headers' y 'records' del método get_attendance_history_by_schedule
        """
        self.data = data

    def export(self):
        raise NotImplementedError("Subclasses must implement export()")


class ExcelAttendanceExporter(AttendanceExporter):
    """Exportador de asistencias a formato Excel"""

    def export(self, schedule_id=None):
        # Crear DataFrame con los registros
        df = pd.DataFrame(self.data['records'])

        if df.empty:
            return self._create_empty_response()

        # Aplicar mapeo de estados
        df['status'] = df['status'].map(self.STATUS_MAP)

        # Formatear fechas
        df['fecha'] = pd.to_datetime(df['class_session__actual_start_time']).dt.strftime('%d-%m-%Y')

        # Pivot para crear la estructura de tabla
        pivot_df = df.pivot(
            index='complete_name',
            columns='fecha',
            values='status'
        ).fillna('Ausente')

        # Resetear índice para tener el nombre como columna
        pivot_df.reset_index(inplace=True)
        pivot_df.rename(columns={'complete_name': 'Nombre del alumno'}, inplace=True)

        # Crear respuesta Excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        filename = self._generate_filename(schedule_id)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Escribir a Excel con formato
        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            pivot_df.to_excel(writer, index=False, sheet_name='Asistencias')

            # Aplicar estilos
            self._apply_styles(writer.sheets['Asistencias'])

        return response

    def _apply_styles(self, worksheet):
        """Aplica estilos al worksheet"""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center_alignment = Alignment(horizontal="center", vertical="center")

        # Aplicar estilos a headers
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        # Ajustar ancho de columnas
        worksheet.column_dimensions['A'].width = 30
        for col in worksheet.iter_cols(min_col=2, max_col=worksheet.max_column):
            worksheet.column_dimensions[col[0].column_letter].width = 15

    def _generate_filename(self, schedule_id=None):
        """Genera el nombre del archivo"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if schedule_id:
            return f'asistencias_schedule_{schedule_id}_{timestamp}.xlsx'
        return f'asistencias_{timestamp}.xlsx'

    def _create_empty_response(self):
        """Crea una respuesta vacía si no hay datos"""
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="asistencias_sin_datos.xlsx"'

        wb = pd.ExcelWriter(response, engine='openpyxl')
        pd.DataFrame({'Mensaje': ['No hay datos de asistencia']}).to_excel(wb, index=False)
        wb.close()

        return response


class CSVAttendanceExporter(AttendanceExporter):
    """Exportador de asistencias a formato CSV"""

    def export(self, schedule_id=None):
        # Crear DataFrame con los registros
        df = pd.DataFrame(self.data['records'])

        if df.empty:
            return self._create_empty_response()

        # Aplicar mapeo de estados
        df['status'] = df['status'].map(self.STATUS_MAP)

        # Formatear fechas
        df['fecha'] = pd.to_datetime(df['class_session__actual_start_time']).dt.strftime('%d-%m-%Y')

        # Pivot para crear la estructura de tabla
        pivot_df = df.pivot(
            index='complete_name',
            columns='fecha',
            values='status'
        ).fillna('Ausente')

        # Resetear índice
        pivot_df.reset_index(inplace=True)
        pivot_df.rename(columns={'complete_name': 'Nombre del alumno'}, inplace=True)

        # Crear respuesta CSV
        response = HttpResponse(content_type='text/csv')
        filename = self._generate_filename(schedule_id)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Escribir CSV
        pivot_df.to_csv(response, index=False, encoding='utf-8-sig')

        return response

    def _generate_filename(self, schedule_id=None):
        """Genera el nombre del archivo"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if schedule_id:
            return f'asistencias_schedule_{schedule_id}_{timestamp}.csv'
        return f'asistencias_{timestamp}.csv'

    def _create_empty_response(self):
        """Crea una respuesta vacía si no hay datos"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="asistencias_sin_datos.csv"'
        response.write('Mensaje\nNo hay datos de asistencia')
        return response


class PDFAttendanceExporter(AttendanceExporter):
    """Exportador de asistencias a formato PDF"""

    def export(self, schedule_id=None):
        # Crear DataFrame con los registros
        df = pd.DataFrame(self.data['records'])

        if df.empty:
            return self._create_empty_response()

        # Aplicar mapeo de estados
        df['status'] = df['status'].map(self.STATUS_MAP)

        # Formatear fechas
        df['fecha'] = pd.to_datetime(df['class_session__actual_start_time']).dt.strftime('%d-%m-%Y')

        # Pivot para crear la estructura de tabla
        pivot_df = df.pivot(
            index='complete_name',
            columns='fecha',
            values='status'
        ).fillna('Ausente')

        # Resetear índice
        pivot_df.reset_index(inplace=True)
        pivot_df.rename(columns={'complete_name': 'Nombre del alumno'}, inplace=True)

        # Ordenar columnas de fechas cronológicamente
        date_columns = [col for col in pivot_df.columns if col != 'Nombre del alumno']
        date_columns_sorted = sorted(date_columns, key=lambda x: datetime.strptime(x, '%d-%m-%Y'))
        
        # Reordenar el DataFrame con las fechas ordenadas
        pivot_df = pivot_df[['Nombre del alumno'] + date_columns_sorted]

        # Crear buffer para el PDF
        buffer = BytesIO()
        
        # Configurar documento (Landscape para tener más espacio horizontal)
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(letter),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=18
        )

        # Contenido del PDF
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center
        )
        timestamp = datetime.now().strftime("%d/%m/%Y %H:%M")
        
        # Dividir columnas en chunks si son muchas (10 fechas por página)
        CHUNK_SIZE = 10
        date_chunks = [date_columns_sorted[i:i + CHUNK_SIZE] for i in range(0, len(date_columns_sorted), CHUNK_SIZE)]
        
        # Si no hay fechas, crear un chunk vacío
        if not date_chunks:
            date_chunks = [[]]
        
        for chunk_idx, date_chunk in enumerate(date_chunks):
            # Agregar título para cada página
            page_title = f"Reporte de Asistencias - Generado el {timestamp}"
            if len(date_chunks) > 1:
                page_title += f" (Página {chunk_idx + 1}/{len(date_chunks)})"
            elements.append(Paragraph(page_title, title_style))
            elements.append(Spacer(1, 12))

            # Seleccionar columnas para este chunk
            chunk_columns = ['Nombre del alumno'] + date_chunk
            chunk_df = pivot_df[chunk_columns]

            # Preparar datos para la tabla
            headers = list(chunk_df.columns)
            data = [headers]
            
            # Filas
            for _, row in chunk_df.iterrows():
                data.append(row.tolist())

            # Crear tabla
            table = Table(data)
            
            # Estilos de la tabla
            style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),  # Alinear nombres a la izquierda
            ])
            
            table.setStyle(style)
            elements.append(table)
            
            # Agregar salto de página si no es el último chunk
            if chunk_idx < len(date_chunks) - 1:
                elements.append(PageBreak())

        # Generar PDF
        doc.build(elements)

        # Preparar respuesta
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        filename = self._generate_filename(schedule_id)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    def _generate_filename(self, schedule_id=None):
        """Genera el nombre del archivo"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if schedule_id:
            return f'asistencias_schedule_{schedule_id}_{timestamp}.pdf'
        return f'asistencias_{timestamp}.pdf'

    def _create_empty_response(self):
        """Crea una respuesta vacía si no hay datos"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        elements.append(Paragraph("No hay datos de asistencia para mostrar.", styles['Normal']))
        doc.build(elements)
        
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="asistencias_sin_datos.pdf"'
        return response